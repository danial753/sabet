import json
import os
import openpyxl
from app import create_app
from app.models import Factory

app = create_app()

def read_excel_list(filepath):
    """خواندن ستون اول فایل اکسل و برگرداندن یک لیست تمیز و یکتا."""
    if not os.path.exists(filepath):
        return []
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    values = set()
    for row in ws.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
        val = row[0]
        if val is not None and str(val).strip():
            values.add(str(val).strip())
    wb.close()
    return sorted(values)

def read_family_structure(filepath):
    """
    خواندن فایل product_families.xlsx با ساختار ستونی:
    ردیف اول هر ستون = نام گروه (دسته)
    از ردیف دوم به بعد = قطعات آن گروه
    """
    if not os.path.exists(filepath):
        return {'families': [], 'products_by_family': {}, 'all_products': []}

    # برای پشتیبانی از iter_cols، read_only=False
    wb = openpyxl.load_workbook(filepath, read_only=False)
    ws = wb.active

    families = []
    products_by_family = {}
    all_products_set = set()

    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, values_only=True):
        family_name = None
        products = []
        for i, cell in enumerate(col):
            val = str(cell).strip() if cell is not None else ''
            if not val:
                continue
            if i == 0:
                family_name = val   # اولین ردیف → نام گروه
            else:
                products.append(val)
                all_products_set.add(val)
        if family_name and products:
            families.append(family_name)
            products_by_family[family_name] = sorted(set(products))

    wb.close()
    return {
        'families': sorted(set(families)),
        'products_by_family': products_by_family,
        'all_products': sorted(all_products_set)
    }

with app.app_context():
    factory_names = [f.name for f in Factory.query.all()]
    if not factory_names:
        print("❌ هیچ کارخانه‌ای در دیتابیس یافت نشد. ابتدا python create_user.py را اجرا کنید.")
        exit(1)

    for factory_name in factory_names:
        base_data = os.path.join('data', factory_name)
        output_dir = os.path.join('json_data', factory_name)
        os.makedirs(output_dir, exist_ok=True)

        # CNC
        cnc_products = read_excel_list(os.path.join(base_data, 'cnc', 'products.xlsx'))
        cnc_families_data = read_family_structure(os.path.join(base_data, 'cnc', 'product_families.xlsx'))
        cnc = {
            'product_names': cnc_products,
            'families': cnc_families_data['families'],
            'products_by_family': cnc_families_data['products_by_family'],
            'part_sizes': read_excel_list(os.path.join(base_data, 'cnc', 'sizes.xlsx')),
            'machine_codes': read_excel_list(os.path.join(base_data, 'cnc', 'machines.xlsx')),
            'operation_stage_codes': read_excel_list(os.path.join(base_data, 'cnc', 'operations.xlsx')),
        }
        with open(os.path.join(output_dir, 'cnc_options.json'), 'w', encoding='utf-8') as f:
            json.dump(cnc, f, ensure_ascii=False, indent=2)

        # ManAll
        manall_products = read_excel_list(os.path.join(base_data, 'manall', 'products.xlsx'))
        manall_families_data = read_family_structure(os.path.join(base_data, 'manall', 'product_families.xlsx'))
        manall = {
            'product_names': manall_products,
            'families': manall_families_data['families'],
            'products_by_family': manall_families_data['products_by_family'],
            'machine_codes': read_excel_list(os.path.join(base_data, 'manall', 'machines.xlsx')),
            'operation_stage_codes': read_excel_list(os.path.join(base_data, 'manall', 'operations.xlsx')),
            'manual_titles': read_excel_list(os.path.join(base_data, 'manall', 'titles.xlsx')),
        }
        with open(os.path.join(output_dir, 'manall_options.json'), 'w', encoding='utf-8') as f:
            json.dump(manall, f, ensure_ascii=False, indent=2)

        # Stoppage
        stoppage = {
            'machine_codes': read_excel_list(os.path.join(base_data, 'stoppage', 'machines.xlsx')),
            'stop_codes': read_excel_list(os.path.join(base_data, 'stoppage', 'stopcodes.xlsx')),
        }
        with open(os.path.join(output_dir, 'stoppage_options.json'), 'w', encoding='utf-8') as f:
            json.dump(stoppage, f, ensure_ascii=False, indent=2)

        print(f"✅ فایل‌های JSON برای «{factory_name}» به‌روزرسانی شدند.")