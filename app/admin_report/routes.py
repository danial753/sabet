# app/admin_report/routes.py
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from app.models import db, User, ProductionReport, StoppageReport
from datetime import datetime, timezone, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from persiantools.jdatetime import JalaliDate, JalaliDateTime
from datetime import datetime as dt_datetime

admin_report_bp = Blueprint('admin_report', __name__)
IRAN_TZ = timezone(timedelta(hours=3, minutes=30))

@admin_report_bp.route('/general-report', methods=['GET'])
@login_required
def general_report_form():
    if not current_user.is_admin:
        flash('دسترسی غیرمجاز', 'danger')
        return redirect(url_for('reports.dashboard'))
    today = JalaliDate.today().strftime('%Y-%m-%d')
    return render_template('admin/general_report.html', today=today)

@admin_report_bp.route('/general-report/download', methods=['POST'])
@login_required
def general_report_download():
    if not current_user.is_admin:
        flash('دسترسی غیرمجاز', 'danger')
        return redirect(url_for('reports.dashboard'))

    user_code = request.form.get('user_code', '').strip()
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')

    prod_query = ProductionReport.query
    stop_query = StoppageReport.query

    if user_code:
        user = User.query.filter_by(code=user_code).first()
        if not user:
            flash('کاربری با این کد یافت نشد.', 'danger')
            return redirect(url_for('admin_report.general_report_form'))
        prod_query = prod_query.filter_by(user_id=user.id)
        stop_query = stop_query.filter_by(user_id=user.id)

    if start_date:
        try:
            start_dt = JalaliDate.strptime(start_date, '%Y-%m-%d').to_gregorian()
            start_dt = dt_datetime.combine(start_dt, dt_datetime.min.time())
            prod_query = prod_query.filter(ProductionReport.date >= start_dt)
            stop_query = stop_query.filter(StoppageReport.date >= start_dt)
        except ValueError:
            flash('فرمت تاریخ شروع نامعتبر است.', 'danger')
            return redirect(url_for('admin_report.general_report_form'))

    if end_date:
        try:
            end_dt = JalaliDate.strptime(end_date, '%Y-%m-%d').to_gregorian()
            end_dt = dt_datetime.combine(end_dt, dt_datetime.max.time())
            prod_query = prod_query.filter(ProductionReport.date <= end_dt)
            stop_query = stop_query.filter(StoppageReport.date <= end_dt)
        except ValueError:
            flash('فرمت تاریخ پایان نامعتبر است.', 'danger')
            return redirect(url_for('admin_report.general_report_form'))

    cnc_reports = prod_query.filter_by(type='CNC').order_by(ProductionReport.date.asc()).all()
    manall_reports = prod_query.filter_by(type='ManAll').order_by(ProductionReport.date.asc()).all()
    stoppage_reports = stop_query.order_by(StoppageReport.date.asc()).all()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color='FFFFFF', name='B Nazanin', size=11)
    header_fill = PatternFill(start_color='1a3a5c', end_color='1a3a5c', fill_type='solid')
    wrap_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    def add_header(ws, headers):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap_alignment
            cell.border = thin_border
        ws.freeze_panes = 'A2'

    def auto_width(ws):
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 5, 40)

    # ---------- برگه CNC ----------
    ws_cnc = wb.create_sheet('CNC')
    ws_cnc.sheet_view.rightToLeft = True
    cnc_headers = [
        'تاریخ ثبت', 'کد اپراتور', 'اسم قطعه', 'شیفت', 'اسم سرپرست', 'ت. تأیید سرپرست',
        'سایز قطعه', 'کد دستگاه', 'کد مرحله کاری', 'تعداد', 'تعداد اصلاحی', 'Feed (%)',
        'ت. بازرسی', 'ت. انبار', 'ت. تأیید نهایی',
        'زمان مورد انتظار', 'زمان اجرایی', 'اختلاف زمانی'
    ]
    add_header(ws_cnc, cnc_headers)
    row = 2
    for r in cnc_reports:
        ws_cnc.cell(row=row, column=1, value=JalaliDateTime.to_jalali(r.date).strftime('%Y-%m-%d %H:%M')).border = thin_border
        ws_cnc.cell(row=row, column=2, value=r.user.code if r.user else '-').border = thin_border
        ws_cnc.cell(row=row, column=3, value=r.product_name or '-').border = thin_border
        ws_cnc.cell(row=row, column=4, value=r.shift_fa).border = thin_border
        ws_cnc.cell(row=row, column=5, value=r.work_type_approved_by.full_name if r.work_type_approved_by else '-').border = thin_border
        ws_cnc.cell(row=row, column=6, value=JalaliDateTime.to_jalali(r.work_type_approval_date).strftime('%Y-%m-%d %H:%M') if r.work_type_approval_date else '-').border = thin_border
        ws_cnc.cell(row=row, column=7, value=r.part_size or '-').border = thin_border
        ws_cnc.cell(row=row, column=8, value=r.machine_code or '-').border = thin_border
        ws_cnc.cell(row=row, column=9, value=r.operation_stage_code or '-').border = thin_border
        ws_cnc.cell(row=row, column=10, value=r.quantity).border = thin_border
        ws_cnc.cell(row=row, column=11, value=r.inspection_quantity if r.inspection_quantity is not None else '-').border = thin_border
        ws_cnc.cell(row=row, column=12, value=f"{r.feed_percent}%" if r.feed_percent else '-').border = thin_border
        ws_cnc.cell(row=row, column=13, value=JalaliDateTime.to_jalali(r.inspection_date).strftime('%Y-%m-%d %H:%M') if r.inspection_date else '-').border = thin_border
        ws_cnc.cell(row=row, column=14, value=JalaliDateTime.to_jalali(r.warehouse_date).strftime('%Y-%m-%d %H:%M') if r.warehouse_date else '-').border = thin_border
        ws_cnc.cell(row=row, column=15, value=JalaliDateTime.to_jalali(r.approval_date).strftime('%Y-%m-%d %H:%M') if r.approval_date else '-').border = thin_border
        ws_cnc.cell(row=row, column=16, value=r.expected_duration_formatted or '-').border = thin_border
        ws_cnc.cell(row=row, column=17, value=r.duration or '-').border = thin_border
        ws_cnc.cell(row=row, column=18, value=r.time_diff_formatted or '-').border = thin_border
        row += 1
    auto_width(ws_cnc)

    # ---------- برگه Manual ----------
    ws_man = wb.create_sheet('Manual')
    ws_man.sheet_view.rightToLeft = True
    man_headers = [
        'تاریخ ثبت', 'کد اپراتور', 'اسم قطعه', 'شیفت', 'اسم سرپرست', 'ت. تأیید سرپرست',
        'عنوان کار منوال', 'تعداد', 'ت. انبار', 'ت. تأیید نهایی',
        'زمان مورد انتظار', 'زمان اجرایی', 'اختلاف زمانی'
    ]
    add_header(ws_man, man_headers)
    row = 2
    for r in manall_reports:
        ws_man.cell(row=row, column=1, value=JalaliDateTime.to_jalali(r.date).strftime('%Y-%m-%d %H:%M')).border = thin_border
        ws_man.cell(row=row, column=2, value=r.user.code if r.user else '-').border = thin_border
        ws_man.cell(row=row, column=3, value=r.product_name or '-').border = thin_border
        ws_man.cell(row=row, column=4, value=r.shift_fa).border = thin_border
        ws_man.cell(row=row, column=5, value=r.work_type_approved_by.full_name if r.work_type_approved_by else '-').border = thin_border
        ws_man.cell(row=row, column=6, value=JalaliDateTime.to_jalali(r.work_type_approval_date).strftime('%Y-%m-%d %H:%M') if r.work_type_approval_date else '-').border = thin_border
        ws_man.cell(row=row, column=7, value=r.manual_title or '-').border = thin_border
        ws_man.cell(row=row, column=8, value=r.quantity).border = thin_border
        ws_man.cell(row=row, column=9, value=JalaliDateTime.to_jalali(r.warehouse_date).strftime('%Y-%m-%d %H:%M') if r.warehouse_date else '-').border = thin_border
        ws_man.cell(row=row, column=10, value=JalaliDateTime.to_jalali(r.approval_date).strftime('%Y-%m-%d %H:%M') if r.approval_date else '-').border = thin_border
        ws_man.cell(row=row, column=11, value=r.expected_duration_formatted or '-').border = thin_border
        ws_man.cell(row=row, column=12, value=r.duration or '-').border = thin_border
        ws_man.cell(row=row, column=13, value=r.time_diff_formatted or '-').border = thin_border
        row += 1
    auto_width(ws_man)

    # ---------- برگه توقف ----------
    ws_stop = wb.create_sheet('توقف')
    ws_stop.sheet_view.rightToLeft = True
    stop_headers = ['تاریخ ثبت', 'کد اپراتور', 'کد توقف', 'دلیل (اختیاری)',
                    'زمان مورد انتظار توقف', 'زمان اجرایی']
    add_header(ws_stop, stop_headers)
    row = 2
    for r in stoppage_reports:
        user = User.query.get(r.user_id) if r.user_id else None
        ws_stop.cell(row=row, column=1, value=JalaliDateTime.to_jalali(r.date).strftime('%Y-%m-%d %H:%M')).border = thin_border
        ws_stop.cell(row=row, column=2, value=user.code if user else '-').border = thin_border
        ws_stop.cell(row=row, column=3, value=r.stop_code or '-').border = thin_border
        ws_stop.cell(row=row, column=4, value=r.reason or '-').border = thin_border
        ws_stop.cell(row=row, column=5, value=r.expected_duration_formatted or '-').border = thin_border
        ws_stop.cell(row=row, column=6, value=r.duration or '-').border = thin_border
        row += 1
    auto_width(ws_stop)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"general_report_{JalaliDate.today().strftime('%Y%m%d')}.xlsx"
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=filename)