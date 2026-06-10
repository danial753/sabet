# app/admin/routes.py
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import login_required, current_user
from app.models import db, User, ProductionReport, StoppageReport, Factory, SystemLog, ConfigSetting, WorkSession, Notification
from werkzeug.security import generate_password_hash
from sqlalchemy import or_
from datetime import datetime, timezone, timedelta
from datetime import datetime as dt_datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from app.log_utils import log_action
from persiantools.jdatetime import JalaliDate
import subprocess
import sys

admin_bp = Blueprint('admin', __name__)
IRAN_TZ = timezone(timedelta(hours=3, minutes=30))

# ------------------------------------------------------------
#   لیست کاربران (با جستجو، فیلتر و مرتب‌سازی)
# ------------------------------------------------------------
@admin_bp.route('/users')
@login_required
def user_list():
    if not current_user.is_admin:
        flash('دسترسی غیرمجاز', 'danger')
        return redirect(url_for('reports.dashboard'))

    q = request.args.get('q', '').strip()
    role = request.args.get('role', '')
    sort = request.args.get('sort', 'id')
    order = request.args.get('order', 'asc')

    query = User.query

    if q:
        search = f"%{q}%"
        query = query.filter(or_(
            User.first_name.ilike(search),
            User.last_name.ilike(search),
            User.code.ilike(search),
            User.username.ilike(search)
        ))

    if role == 'admin':
        query = query.filter_by(is_admin=True)
    elif role == 'approver':
        query = query.filter_by(is_approver=True)
    elif role == 'shift_planner':
        query = query.filter_by(is_shift_planner=True)
    elif role == 'quality_inspector':
        query = query.filter_by(is_quality_inspector=True)
    elif role == 'warehouse':
        query = query.filter_by(is_warehouse=True)
    elif role == 'operator':
        query = query.filter_by(is_admin=False, is_approver=False, is_shift_planner=False,
                                is_quality_inspector=False, is_warehouse=False)

    if sort in ['id', 'first_name', 'last_name', 'code', 'username']:
        col = getattr(User, sort)
        if order == 'desc':
            query = query.order_by(col.desc())
        else:
            query = query.order_by(col.asc())
    else:
        query = query.order_by(User.id.asc())

    users = query.all()
    return render_template('admin/user_list.html', users=users, q=q, role=role, sort=sort, order=order)

# ------------------------------------------------------------
#   ایجاد کاربر جدید
# ------------------------------------------------------------
@admin_bp.route('/users/create', methods=['GET', 'POST'])
@login_required
def user_create():
    if not current_user.is_admin:
        flash('دسترسی غیرمجاز', 'danger')
        return redirect(url_for('reports.dashboard'))

    factories = Factory.query.all()

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        code = request.form.get('code')
        factory_id = request.form.get('factory_id', type=int)
        is_approver = request.form.get('is_approver') == 'on'
        is_shift_planner = request.form.get('is_shift_planner') == 'on'
        is_quality_inspector = request.form.get('is_quality_inspector') == 'on'
        is_warehouse = request.form.get('is_warehouse') == 'on'

        if not all([username, password, first_name, last_name, code, factory_id]):
            flash('همه فیلدها الزامی هستند.', 'danger')
            return render_template('admin/user_create.html', factories=factories)

        if User.query.filter_by(username=username).first():
            flash('این نام کاربری قبلاً ثبت شده است.', 'danger')
            return render_template('admin/user_create.html', factories=factories)

        if User.query.filter_by(code=code).first():
            flash('این کد قبلاً ثبت شده است.', 'danger')
            return render_template('admin/user_create.html', factories=factories)

        new_user = User(
            username=username,
            password=generate_password_hash(password),
            first_name=first_name,
            last_name=last_name,
            code=code,
            is_admin=False,
            is_approver=is_approver,
            is_shift_planner=is_shift_planner,
            is_quality_inspector=is_quality_inspector,
            is_warehouse=is_warehouse,
            factory_id=factory_id
        )
        db.session.add(new_user)
        db.session.commit()
        log_action(current_user, 'ایجاد کاربر', f'کاربر {username} با کد {code} در کارخانه {new_user.factory.name} ایجاد شد')
        flash('کاربر جدید با موفقیت ایجاد شد.', 'success')
        return redirect(url_for('admin.user_list'))

    return render_template('admin/user_create.html', factories=factories)

# ------------------------------------------------------------
#   ویرایش کاربر
# ------------------------------------------------------------
@admin_bp.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
def user_edit(user_id):
    if not current_user.is_admin:
        flash('دسترسی غیرمجاز', 'danger')
        return redirect(url_for('reports.dashboard'))

    user = User.query.get(user_id)
    if not user:
        flash('کاربر مورد نظر یافت نشد.', 'danger')
        return redirect(url_for('admin.user_list'))

    factories = Factory.query.all()

    if request.method == 'POST':
        username = request.form.get('username')
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        code = request.form.get('code')
        factory_id = request.form.get('factory_id', type=int)
        password = request.form.get('password')
        is_approver = request.form.get('is_approver') == 'on'
        is_shift_planner = request.form.get('is_shift_planner') == 'on'
        is_quality_inspector = request.form.get('is_quality_inspector') == 'on'
        is_warehouse = request.form.get('is_warehouse') == 'on'

        if not all([username, first_name, last_name, code, factory_id]):
            flash('همه فیلدها به جز رمز عبور الزامی هستند.', 'danger')
            return render_template('admin/user_edit.html', user=user, factories=factories)

        if User.query.filter(User.username == username, User.id != user.id).first():
            flash('این نام کاربری قبلاً توسط کاربر دیگری ثبت شده است.', 'danger')
            return render_template('admin/user_edit.html', user=user, factories=factories)
        if User.query.filter(User.code == code, User.id != user.id).first():
            flash('این کد قبلاً توسط کاربر دیگری ثبت شده است.', 'danger')
            return render_template('admin/user_edit.html', user=user, factories=factories)

        user.username = username
        user.first_name = first_name
        user.last_name = last_name
        user.code = code
        user.factory_id = factory_id
        user.is_approver = is_approver
        user.is_shift_planner = is_shift_planner
        user.is_quality_inspector = is_quality_inspector
        user.is_warehouse = is_warehouse
        if password:
            user.password = generate_password_hash(password)
        db.session.commit()
        log_action(current_user, 'ویرایش کاربر', f'کاربر {username} با کد {code} ویرایش شد')
        flash('اطلاعات کاربر با موفقیت به‌روزرسانی شد.', 'success')
        return redirect(url_for('admin.user_list'))

    return render_template('admin/user_edit.html', user=user, factories=factories)

# ------------------------------------------------------------
#   حذف کاربر (تکی)
# ------------------------------------------------------------
@admin_bp.route('/users/<int:user_id>/delete', methods=['POST'])
@login_required
def user_delete(user_id):
    if not current_user.is_admin:
        flash('دسترسی غیرمجاز', 'danger')
        return redirect(url_for('reports.dashboard'))

    user = User.query.get(user_id)
    if not user:
        flash('کاربر مورد نظر یافت نشد.', 'danger')
        return redirect(url_for('admin.user_list'))

    if user.id == current_user.id:
        flash('نمی‌توانید حساب کاربری خود را حذف کنید.', 'danger')
        return redirect(url_for('admin.user_list'))

    username = user.username
    ProductionReport.query.filter_by(user_id=user.id).delete()
    StoppageReport.query.filter_by(user_id=user.id).delete()
    WorkSession.query.filter_by(user_id=user.id).delete()
    db.session.delete(user)
    db.session.commit()
    log_action(current_user, 'حذف کاربر', f'کاربر {username} حذف شد')
    flash('کاربر با موفقیت حذف شد.', 'success')
    return redirect(url_for('admin.user_list'))

# ------------------------------------------------------------
#   حذف دسته‌جمعی کاربران
# ------------------------------------------------------------
@admin_bp.route('/users/delete-multiple', methods=['POST'])
@login_required
def user_delete_multiple():
    if not current_user.is_admin:
        flash('دسترسی غیرمجاز', 'danger')
        return redirect(url_for('reports.dashboard'))

    user_ids = request.form.getlist('user_ids')
    if not user_ids:
        flash('هیچ کاربری انتخاب نشده است.', 'warning')
        return redirect(url_for('admin.user_list'))

    try:
        ids = [int(uid) for uid in user_ids]
    except ValueError:
        flash('شناسه‌های نامعتبر', 'danger')
        return redirect(url_for('admin.user_list'))

    users_to_delete = User.query.filter(User.id.in_(ids), User.id != current_user.id).all()
    names = [u.username for u in users_to_delete]
    for user in users_to_delete:
        ProductionReport.query.filter_by(user_id=user.id).delete()
        StoppageReport.query.filter_by(user_id=user.id).delete()
        WorkSession.query.filter_by(user_id=user.id).delete()
        db.session.delete(user)

    db.session.commit()
    log_action(current_user, 'حذف دسته‌جمعی کاربران', f'{len(users_to_delete)} کاربر حذف شدند: {", ".join(names)}')
    flash(f'{len(users_to_delete)} کاربر با موفقیت حذف شد.', 'success')
    return redirect(url_for('admin.user_list'))

# ------------------------------------------------------------
#   مشاهده گزارش‌های یک کاربر (با تاریخ شمسی)
# ------------------------------------------------------------
@admin_bp.route('/view-reports', methods=['GET', 'POST'])
@login_required
def view_user_reports():
    if not current_user.is_admin:
        flash('دسترسی غیرمجاز', 'danger')
        return redirect(url_for('reports.dashboard'))

    user = None
    cnc_reports = []
    manall_reports = []
    stoppage_reports = []
    start_date = end_date = None

    if request.method == 'POST':
        code = request.form.get('user_code')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')

        if code:
            user = User.query.filter_by(code=code).first()
            if not user:
                flash('کاربری با این کد یافت نشد.', 'danger')
            else:
                prod_base = ProductionReport.query.filter_by(user_id=user.id)
                stop_base = StoppageReport.query.filter_by(user_id=user.id)

                if start_date:
                    try:
                        start_dt = JalaliDate.strptime(start_date, '%Y-%m-%d').to_gregorian()
                        start_dt = dt_datetime.combine(start_dt, dt_datetime.min.time())
                        prod_base = prod_base.filter(ProductionReport.date >= start_dt)
                        stop_base = stop_base.filter(StoppageReport.date >= start_dt)
                    except ValueError:
                        flash('فرمت تاریخ شروع نامعتبر است (YYYY-MM-DD).', 'danger')
                if end_date:
                    try:
                        end_dt = JalaliDate.strptime(end_date, '%Y-%m-%d').to_gregorian()
                        end_dt = dt_datetime.combine(end_dt, dt_datetime.max.time())
                        prod_base = prod_base.filter(ProductionReport.date <= end_dt)
                        stop_base = stop_base.filter(StoppageReport.date <= end_dt)
                    except ValueError:
                        flash('فرمت تاریخ پایان نامعتبر است (YYYY-MM-DD).', 'danger')

                cnc_reports = prod_base.filter_by(type='CNC').order_by(ProductionReport.date.desc()).all()
                manall_reports = prod_base.filter_by(type='ManAll').order_by(ProductionReport.date.desc()).all()
                stoppage_reports = stop_base.order_by(StoppageReport.date.desc()).all()
        else:
            flash('لطفاً کد کاربر را وارد کنید.', 'warning')

    return render_template('admin/view_user_reports.html',
                           user=user,
                           cnc_reports=cnc_reports,
                           manall_reports=manall_reports,
                           stoppage_reports=stoppage_reports,
                           start_date=start_date,
                           end_date=end_date)

# ------------------------------------------------------------
#   دانلود فایل اکسل از گزارش‌های کاربر (با اطلاعات انبار)
# ------------------------------------------------------------
@admin_bp.route('/download-reports-excel', methods=['POST'])
@login_required
def download_user_reports_excel():
    if not current_user.is_admin:
        flash('دسترسی غیرمجاز', 'danger')
        return redirect(url_for('reports.dashboard'))

    code = request.form.get('user_code')
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')

    user = User.query.filter_by(code=code).first() if code else None
    if not user:
        flash('کاربری با این کد یافت نشد.', 'danger')
        return redirect(url_for('admin.view_user_reports'))

    prod_base = ProductionReport.query.filter_by(user_id=user.id)
    stop_base = StoppageReport.query.filter_by(user_id=user.id)

    if start_date:
        try:
            start_dt = JalaliDate.strptime(start_date, '%Y-%m-%d').to_gregorian()
            start_dt = dt_datetime.combine(start_dt, dt_datetime.min.time())
            prod_base = prod_base.filter(ProductionReport.date >= start_dt)
            stop_base = stop_base.filter(StoppageReport.date >= start_dt)
        except ValueError:
            pass
    if end_date:
        try:
            end_dt = JalaliDate.strptime(end_date, '%Y-%m-%d').to_gregorian()
            end_dt = dt_datetime.combine(end_dt, dt_datetime.max.time())
            prod_base = prod_base.filter(ProductionReport.date <= end_dt)
            stop_base = stop_base.filter(StoppageReport.date <= end_dt)
        except ValueError:
            pass

    cnc_reports = prod_base.filter_by(type='CNC').order_by(ProductionReport.date.desc()).all()
    manall_reports = prod_base.filter_by(type='ManAll').order_by(ProductionReport.date.desc()).all()
    stoppage_reports = stop_base.order_by(StoppageReport.date.desc()).all()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='343a40', end_color='343a40', fill_type='solid')
    wrap_alignment = Alignment(wrap_text=True)

    def add_header(ws, headers):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap_alignment

    # برگه اطلاعات کاربر
    ws_info = wb.create_sheet('اطلاعات کاربر')
    info_data = [
        ('کد کاربری', user.code),
        ('نام', user.first_name),
        ('نام خانوادگی', user.last_name),
        ('نام کامل', user.full_name),
        ('نام کاربری', user.username),
        ('کارخانه', user.factory.name if user.factory else '-'),
        ('ادمین', 'بله' if user.is_admin else 'خیر'),
        ('تأییدکننده کیفیت', 'بله' if user.is_approver else 'خیر'),
        ('سرپرست تولید', 'بله' if user.is_shift_planner else 'خیر'),
        ('بازرس کیفیت', 'بله' if user.is_quality_inspector else 'خیر'),
        ('انبار', 'بله' if user.is_warehouse else 'خیر'),
        ('تاریخ گزارش', datetime.now(IRAN_TZ).strftime('%Y-%m-%d %H:%M'))
    ]
    for i, (key, value) in enumerate(info_data, 1):
        ws_info.cell(row=i, column=1, value=key).font = Font(bold=True)
        ws_info.cell(row=i, column=2, value=value)
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 30

    # برگه CNC
    ws_cnc = wb.create_sheet('CNC')
    headers_cnc = ['تاریخ ثبت', 'شیفت', 'نوع کار', 'قطعه', 'سایز', 'کد دستگاه', 'مرحله کاری', 'تعداد (اپراتور)', 'تعداد تأییدشده', 'تعداد انبار', 'شروع', 'پایان', 'مدت واقعی', 'زمان مورد انتظار']
    add_header(ws_cnc, headers_cnc)
    for i, r in enumerate(cnc_reports, 2):
        ws_cnc.cell(row=i, column=1, value=r.date.strftime('%Y-%m-%d %H:%M'))
        ws_cnc.cell(row=i, column=2, value=r.shift_fa)
        ws_cnc.cell(row=i, column=3, value=r.work_type_fa)
        ws_cnc.cell(row=i, column=4, value=r.product_name)
        ws_cnc.cell(row=i, column=5, value=r.part_size)
        ws_cnc.cell(row=i, column=6, value=r.machine_code)
        ws_cnc.cell(row=i, column=7, value=r.operation_stage_code)
        ws_cnc.cell(row=i, column=8, value=r.quantity)
        ws_cnc.cell(row=i, column=9, value=r.approved_quantity if r.is_approved else '')
        ws_cnc.cell(row=i, column=10, value=r.warehouse_quantity if r.warehouse_quantity is not None else '')
        ws_cnc.cell(row=i, column=11, value=r.start_time.strftime('%Y-%m-%d %H:%M:%S') if r.start_time else '')
        ws_cnc.cell(row=i, column=12, value=r.end_time.strftime('%Y-%m-%d %H:%M:%S') if r.end_time else '')
        ws_cnc.cell(row=i, column=13, value=r.duration or '')
        ws_cnc.cell(row=i, column=14, value=r.expected_duration_formatted or '')

    # برگه Manual
    ws_man = wb.create_sheet('Manual')
    headers_man = ['تاریخ ثبت', 'شیفت', 'نوع کار', 'قطعه', 'کد دستگاه', 'مرحله کاری', 'عنوان کار', 'تعداد (اپراتور)', 'تعداد تأییدشده', 'تعداد انبار', 'شروع', 'پایان', 'مدت واقعی', 'زمان مورد انتظار']
    add_header(ws_man, headers_man)
    for i, r in enumerate(manall_reports, 2):
        ws_man.cell(row=i, column=1, value=r.date.strftime('%Y-%m-%d %H:%M'))
        ws_man.cell(row=i, column=2, value=r.shift_fa)
        ws_man.cell(row=i, column=3, value=r.work_type_fa)
        ws_man.cell(row=i, column=4, value=r.product_name)
        ws_man.cell(row=i, column=5, value=r.machine_code)
        ws_man.cell(row=i, column=6, value=r.operation_stage_code)
        ws_man.cell(row=i, column=7, value=r.manual_title or '')
        ws_man.cell(row=i, column=8, value=r.quantity)
        ws_man.cell(row=i, column=9, value=r.approved_quantity if r.is_approved else '')
        ws_man.cell(row=i, column=10, value=r.warehouse_quantity if r.warehouse_quantity is not None else '')
        ws_man.cell(row=i, column=11, value=r.start_time.strftime('%Y-%m-%d %H:%M:%S') if r.start_time else '')
        ws_man.cell(row=i, column=12, value=r.end_time.strftime('%Y-%m-%d %H:%M:%S') if r.end_time else '')
        ws_man.cell(row=i, column=13, value=r.duration or '')
        ws_man.cell(row=i, column=14, value=r.expected_duration_formatted or '')

    # برگه توقف
    ws_stop = wb.create_sheet('توقف')
    headers_stop = ['تاریخ ثبت', 'کد دستگاه', 'کد توقف', 'دلیل', 'شروع', 'پایان (راه‌اندازی)', 'مدت واقعی', 'زمان مورد انتظار', 'اختلاف']
    add_header(ws_stop, headers_stop)
    for i, r in enumerate(stoppage_reports, 2):
        ws_stop.cell(row=i, column=1, value=r.date.strftime('%Y-%m-%d %H:%M'))
        ws_stop.cell(row=i, column=2, value=r.machine_code)
        ws_stop.cell(row=i, column=3, value=r.stop_code)
        ws_stop.cell(row=i, column=4, value=r.reason or '')
        ws_stop.cell(row=i, column=5, value=r.start_time.strftime('%Y-%m-%d %H:%M:%S') if r.start_time else '')
        ws_stop.cell(row=i, column=6, value=r.end_time.strftime('%Y-%m-%d %H:%M:%S') if r.end_time else '')
        ws_stop.cell(row=i, column=7, value=r.duration or '')
        ws_stop.cell(row=i, column=8, value=r.expected_duration_formatted or '')
        ws_stop.cell(row=i, column=9, value=r.time_diff_formatted or '')

    for ws in [ws_info, ws_cnc, ws_man, ws_stop]:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 5, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reports_{user.code}_{user.last_name}_{datetime.now(IRAN_TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"
    log_action(current_user, 'دانلود گزارش کاربر', f'گزارش کاربر {user.code} دانلود شد')
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=filename)

# ------------------------------------------------------------
#   گزارش پیشرفته (فیلترهای ترکیبی با تاریخ شمسی)
# ------------------------------------------------------------
@admin_bp.route('/advanced-report', methods=['GET', 'POST'])
@login_required
def advanced_report():
    if not current_user.is_admin:
        flash('دسترسی غیرمجاز', 'danger')
        return redirect(url_for('reports.dashboard'))

    machines = [r[0] for r in db.session.query(ProductionReport.machine_code).distinct()]
    products = [r[0] for r in db.session.query(ProductionReport.product_name).distinct()]
    reports = []

    if request.method == 'POST':
        machine = request.form.get('machine')
        product = request.form.get('product')
        shift = request.form.get('shift')
        work_type = request.form.get('work_type')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')

        query = ProductionReport.query
        if machine:
            query = query.filter_by(machine_code=machine)
        if product:
            query = query.filter_by(product_name=product)
        if shift:
            query = query.filter_by(shift=shift)
        if work_type:
            query = query.filter_by(work_type=work_type)
        if start_date:
            try:
                start_dt = JalaliDate.strptime(start_date, '%Y-%m-%d').to_gregorian()
                start_dt = dt_datetime.combine(start_dt, dt_datetime.min.time())
                query = query.filter(ProductionReport.date >= start_dt)
            except ValueError:
                flash('فرمت تاریخ شروع نامعتبر است.', 'danger')
        if end_date:
            try:
                end_dt = JalaliDate.strptime(end_date, '%Y-%m-%d').to_gregorian()
                end_dt = dt_datetime.combine(end_dt, dt_datetime.max.time())
                query = query.filter(ProductionReport.date <= end_dt)
            except ValueError:
                flash('فرمت تاریخ پایان نامعتبر است.', 'danger')
        reports = query.order_by(ProductionReport.date.desc()).all()

    return render_template('admin/advanced_report.html',
                           machines=machines,
                           products=products,
                           reports=reports)

# ------------------------------------------------------------
#   ویرایش گزارش تولید (ادمین) - کامل با تمام فیلدها
# ------------------------------------------------------------
@admin_bp.route('/edit-production/<int:report_id>', methods=['GET', 'POST'])
@login_required
def edit_production_report(report_id):
    if not current_user.is_admin:
        flash('دسترسی غیرمجاز', 'danger')
        return redirect(url_for('reports.dashboard'))

    report = ProductionReport.query.get(report_id)
    if not report:
        flash('گزارش یافت نشد.', 'danger')
        return redirect(url_for('admin.view_user_reports'))

    operators = User.query.filter_by(is_admin=False, is_approver=False, is_shift_planner=False,
                                     is_quality_inspector=False, is_warehouse=False).all()

    if request.method == 'POST':
        try:
            user_id = int(request.form.get('user_id'))
            report.user_id = user_id

            report.product_name = request.form.get('product_name', '').strip()
            report.shift = request.form.get('shift', '')
            report.work_type = request.form.get('work_type', '')
            report.machine_code = request.form.get('machine_code', '').strip()
            report.operation_stage_code = request.form.get('operation_stage_code', '').strip()

            if report.type == 'CNC':
                report.part_size = request.form.get('part_size', '').strip()
            if report.type == 'ManAll':
                report.manual_title = request.form.get('manual_title', '').strip()

            report.quantity = int(request.form.get('quantity', 0))

            approved_quantity = request.form.get('approved_quantity')
            if approved_quantity and approved_quantity.strip():
                report.approved_quantity = int(approved_quantity)
                report.is_approved = True
                if not report.approved_by_id:
                    report.approved_by_id = current_user.id
                    report.approval_date = datetime.now(IRAN_TZ)
            else:
                report.approved_quantity = None
                report.is_approved = False
                report.approved_by_id = None
                report.approval_date = None

            start_time_str = request.form.get('start_time')
            end_time_str = request.form.get('end_time')
            if start_time_str:
                report.start_time = datetime.strptime(start_time_str, '%Y-%m-%dT%H:%M')
            if end_time_str:
                report.end_time = datetime.strptime(end_time_str, '%Y-%m-%dT%H:%M')

            if report.work_type and not report.work_type_approved:
                report.work_type_approved = True
                report.work_type_approved_by_id = current_user.id
                report.work_type_approval_date = datetime.now(IRAN_TZ)

            db.session.commit()
            log_action(current_user, 'ویرایش کامل گزارش تولید', f'گزارش {report.id} اپراتور {report.user.code} ویرایش شد')
            flash('گزارش با موفقیت ویرایش شد.', 'success')
            return redirect(url_for('admin.view_user_reports'))
        except Exception as e:
            flash(f'خطا در ذخیره تغییرات: {str(e)}', 'danger')

    return render_template('admin/edit_production_report.html', report=report, operators=operators)

# ------------------------------------------------------------
#   ویرایش گزارش توقف (ادمین) - کامل
# ------------------------------------------------------------
@admin_bp.route('/edit-stoppage/<int:report_id>', methods=['GET', 'POST'])
@login_required
def edit_stoppage_report(report_id):
    if not current_user.is_admin:
        flash('دسترسی غیرمجاز', 'danger')
        return redirect(url_for('reports.dashboard'))

    report = StoppageReport.query.get(report_id)
    if not report:
        flash('گزارش یافت نشد.', 'danger')
        return redirect(url_for('admin.view_user_reports'))

    if request.method == 'POST':
        try:
            report.machine_code = request.form.get('machine_code', '').strip()
            report.stop_code = request.form.get('stop_code', '').strip()
            report.reason = request.form.get('reason', '').strip()

            start_time_str = request.form.get('start_time')
            end_time_str = request.form.get('end_time')

            if start_time_str:
                report.start_time = datetime.strptime(start_time_str, '%Y-%m-%dT%H:%M')
            if end_time_str:
                report.end_time = datetime.strptime(end_time_str, '%Y-%m-%dT%H:%M')

            db.session.commit()
            log_action(current_user, 'ویرایش کامل گزارش توقف', f'گزارش توقف {report.id} کاربر {report.user_id} ویرایش شد')
            flash('گزارش توقف با موفقیت ویرایش شد.', 'success')
            return redirect(url_for('admin.view_user_reports'))
        except Exception as e:
            flash(f'خطا در ذخیره تغییرات: {str(e)}', 'danger')

    return render_template('admin/edit_stoppage_report.html', report=report)

# ------------------------------------------------------------
#   گزارش ساعات کاری اپراتورها
# ------------------------------------------------------------
@admin_bp.route('/operator-work-time', methods=['GET', 'POST'])
@login_required
def operator_work_time():
    if not current_user.is_admin:
        flash('دسترسی غیرمجاز', 'danger')
        return redirect(url_for('reports.dashboard'))

    operator_id = request.args.get('operator_id', type=int)
    start_date = request.form.get('start_date') or request.args.get('start_date')
    end_date = request.form.get('end_date') or request.args.get('end_date')

    query = WorkSession.query
    if operator_id:
        query = query.filter_by(user_id=operator_id)

    if start_date:
        try:
            s_dt = JalaliDate.strptime(start_date, '%Y-%m-%d').to_gregorian()
            s_dt = dt_datetime.combine(s_dt, dt_datetime.min.time())
            query = query.filter(WorkSession.login_time >= s_dt)
        except ValueError:
            pass
    if end_date:
        try:
            e_dt = JalaliDate.strptime(end_date, '%Y-%m-%d').to_gregorian()
            e_dt = dt_datetime.combine(e_dt, dt_datetime.max.time())
            query = query.filter(WorkSession.login_time <= e_dt)
        except ValueError:
            pass

    sessions = query.order_by(WorkSession.login_time.desc()).all()

    total_minutes = 0
    for s in sessions:
        if s.duration_minutes:
            total_minutes += s.duration_minutes
    total_hours = round(total_minutes / 60, 1)

    operators = User.query.filter_by(is_operator=True).order_by(User.last_name).all()

    return render_template('admin/operator_work_time.html',
                           sessions=sessions,
                           operators=operators,
                           operator_id=operator_id,
                           start_date=start_date,
                           end_date=end_date,
                           total_hours=total_hours)

# ------------------------------------------------------------
#   گزارش هشدارها (ادمین)
# ------------------------------------------------------------
@admin_bp.route('/warning-reports', methods=['GET', 'POST'])
@login_required
def warning_reports():
    if not current_user.is_admin:
        flash('دسترسی غیرمجاز', 'danger')
        return redirect(url_for('reports.dashboard'))

    # پارامترها
    start_date = request.form.get('start_date') or request.args.get('start_date')
    end_date = request.form.get('end_date') or request.args.get('end_date')
    operator_id = request.args.get('operator_id', type=int)
    machine = request.args.get('machine')
    product = request.args.get('product')

    # کوئری‌های پایه
    prod_query = ProductionReport.query.filter(ProductionReport.end_time != None)
    stop_query = StoppageReport.query.filter(StoppageReport.end_time != None)

    if operator_id:
        prod_query = prod_query.filter(ProductionReport.user_id == operator_id)
        stop_query = stop_query.filter(StoppageReport.user_id == operator_id)

    if machine:
        prod_query = prod_query.filter(ProductionReport.machine_code == machine)
        stop_query = stop_query.filter(StoppageReport.machine_code == machine)

    if product:
        prod_query = prod_query.filter(ProductionReport.product_name == product)

    # فیلتر تاریخ
    if start_date:
        try:
            s_dt = JalaliDate.strptime(start_date, '%Y-%m-%d').to_gregorian()
            s_dt = dt_datetime.combine(s_dt, dt_datetime.min.time())
            prod_query = prod_query.filter(ProductionReport.date >= s_dt)
            stop_query = stop_query.filter(StoppageReport.date >= s_dt)
        except ValueError:
            pass
    if end_date:
        try:
            e_dt = JalaliDate.strptime(end_date, '%Y-%m-%d').to_gregorian()
            e_dt = dt_datetime.combine(e_dt, dt_datetime.max.time())
            prod_query = prod_query.filter(ProductionReport.date <= e_dt)
            stop_query = stop_query.filter(StoppageReport.date <= e_dt)
        except ValueError:
            pass

    # دریافت گزارش‌ها
    prod_reports = prod_query.order_by(ProductionReport.date.desc()).all()
    stop_reports = stop_query.order_by(StoppageReport.date.desc()).all()

    # فیلتر بر اساس duration_warning (property)
    warning_prods = [r for r in prod_reports if r.duration_warning]
    warning_stops = [r for r in stop_reports if r.duration_warning]

    # attach user_obj به توقف‌ها
    for r in warning_stops:
        r.user_obj = User.query.get(r.user_id) if r.user_id else None

    # لیست‌های فیلتر
    operators = User.query.filter_by(is_operator=True).order_by(User.last_name).all()
    machines = [r[0] for r in db.session.query(ProductionReport.machine_code).distinct()]
    products = [r[0] for r in db.session.query(ProductionReport.product_name).distinct()]

    # خلاصه
    total_warnings = len(warning_prods) + len(warning_stops)
    avg_deviation = 0
    if total_warnings > 0:
        total_dev = 0
        for r in warning_prods:
            if r.expected_duration_seconds and r.duration_seconds:
                total_dev += (r.duration_seconds - r.expected_duration_seconds) / r.expected_duration_seconds * 100
        for r in warning_stops:
            if r.expected_duration_seconds and r.duration_seconds:
                total_dev += (r.duration_seconds - r.expected_duration_seconds) / r.expected_duration_seconds * 100
        avg_deviation = round(total_dev / total_warnings, 1)

    return render_template('admin/warning_reports.html',
                           warning_prods=warning_prods,
                           warning_stops=warning_stops,
                           operators=operators,
                           machines=machines,
                           products=products,
                           operator_id=operator_id,
                           start_date=start_date,
                           end_date=end_date,
                           total_warnings=total_warnings,
                           avg_deviation=avg_deviation)

# ------------------------------------------------------------
#   API اعلان‌ها (برای زنگوله سرپرست)
# ------------------------------------------------------------
@admin_bp.route('/notifications')
@login_required
def get_notifications():
    if not current_user.is_admin and not current_user.is_shift_planner:
        return jsonify([])
    notifs = Notification.query.order_by(Notification.created_at.desc()).limit(50).all()
    return jsonify([{
        'id': n.id,
        'message': n.message,
        'report_id': n.report_id,
        'operator_name': n.operator_name,
        'product_name': n.product_name,
        'quantity': n.quantity,
        'actual_duration': n.actual_duration,
        'is_read': n.is_read,
        'created_at': n.created_at.strftime('%Y-%m-%d %H:%M')
    } for n in notifs])

@admin_bp.route('/notifications/read/<int:id>', methods=['POST'])
@login_required
def mark_notification_read(id):
    if not current_user.is_admin and not current_user.is_shift_planner:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    notif = Notification.query.get(id)
    if notif:
        notif.is_read = True
        db.session.commit()
    return jsonify({'success': True})

# ------------------------------------------------------------
#   لاگ سیستم
# ------------------------------------------------------------
@admin_bp.route('/system-logs')
@login_required
def system_logs():
    if not current_user.is_admin:
        flash('دسترسی غیرمجاز', 'danger')
        return redirect(url_for('reports.dashboard'))
    page = request.args.get('page', 1, type=int)
    logs = SystemLog.query.order_by(SystemLog.timestamp.desc()).paginate(page=page, per_page=50)
    return render_template('admin/system_logs.html', logs=logs)

# ------------------------------------------------------------
#   پیکربندی (تنظیمات)
# ------------------------------------------------------------
@admin_bp.route('/config', methods=['GET', 'POST'])
@login_required
def config():
    if not current_user.is_admin:
        flash('دسترسی غیرمجاز', 'danger')
        return redirect(url_for('reports.dashboard'))

    if request.method == 'POST':
        warning_threshold = request.form.get('warning_threshold')
        if warning_threshold:
            setting = ConfigSetting.query.filter_by(key='warning_threshold').first()
            if setting:
                setting.value = warning_threshold
            else:
                db.session.add(ConfigSetting(key='warning_threshold', value=warning_threshold))
            db.session.commit()
            log_action(current_user, 'تغییر تنظیمات', f'حد آستانه هشدار به {warning_threshold}% تغییر یافت')
            flash('تنظیمات ذخیره شد.', 'success')
        return redirect(url_for('admin.config'))

    setting = ConfigSetting.query.filter_by(key='warning_threshold').first()
    current_threshold = setting.value if setting else '5'
    return render_template('admin/config.html', current_threshold=current_threshold)

# ------------------------------------------------------------
#   به‌روزرسانی فایل‌های JSON از روی اکسل
# ------------------------------------------------------------
@admin_bp.route('/update-json', methods=['POST'])
@login_required
def update_json_data():
    if not current_user.is_admin:
        flash('دسترسی غیرمجاز', 'danger')
        return redirect(url_for('reports.dashboard'))

    try:
        result = subprocess.run([sys.executable, 'convert_excel_to_json.py'],
                                capture_output=True, text=True, timeout=30)
        if result.returncode == 0:
            flash('فایل‌های JSON با موفقیت به‌روزرسانی شدند.', 'success')
        else:
            flash(f'خطا در به‌روزرسانی: {result.stderr}', 'danger')
    except Exception as e:
        flash(f'خطای سیستمی: {str(e)}', 'danger')

    return redirect(url_for('admin.config'))